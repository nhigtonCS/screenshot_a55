#--------------------------
#Import required modules

#Factory modules
from io import BytesIO
from datetime import date
import subprocess
import os
import os.path
import shutil
import glob
import time
import re

#3rd party modules
from openpyxl import Workbook, load_workbook, drawing
from openpyxl.drawing import line, image
from openpyxl.drawing.image import Image
from PIL import Image
import easygui
from easygui import *
import selenium
import getpass
import cv2
import numpy as np
import choco
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import chromedriver_autoinstaller

#----------------------------------------------------------------

#Pole class to hold BNG Grid References for a specific pole.
class Pole:
    def __init__(BT_ID,xGRF,yGRF,AFN1,AFN2):

        BT_ID.xGRF = xGRF
        BT_ID.yGRF = yGRF
        BT_ID.AFN1 = AFN1
        BT_ID.AFN2 = AFN2

#
FILEBROWSER_PATH = os.path.join(os.getenv('WINDIR'), 'explorer.exe')

def explore(path):
    # explorer would choke on forward slashes
    path = os.path.normpath(path)

    if os.path.isdir(path):
        subprocess.run([FILEBROWSER_PATH, path])
    elif os.path.isfile(path):
        subprocess.run([FILEBROWSER_PATH, '/select,', os.path.normpath(path)])

user = getpass.getuser() #get windows username

#config = configparser.ConfigParser()

chromedriver_autoinstaller.install()

print("Select folder containing A55 documents. Note: please make sure folder only contains A55 files that you wish to edit.")

A55Path = easygui.diropenbox()

print("Select folder to contain your screenshots")

screenshotPath = easygui.diropenbox()

driver = webdriver.Chrome()
        
wait = WebDriverWait(driver, 600)
        
# Open the website
driver.get('https://www.beta.openreach.co.uk/cpportal/login')

url = 'https://www.beta.openreach.co.uk/cpportal/dashboard'
        
wait.until(EC.url_matches(url))
        


for filename in os.listdir(A55Path):
    if filename.endswith(".xlsx"):

        driver.get("https://www.beta.openreach.co.uk/ormaps/pia/v2/")
        
        #--------------------------------------------
        #Get grid reference for relevant A55 element
        substring = "A55"

        A55Book = load_workbook(A55Path + '\\' + filename)

        #check containing A55 text
        for s in range(len(A55Book.sheetnames)):
            if substring in A55Book.sheetnames[s]:
                break
        A55Book.active = s

        cover = A55Book.active

        fullGRF = cover['O4'].internal_value

        GRF = fullGRF.split(",")
        GRF[0] = GRF[0].strip()
        GRF[1] = GRF[1].strip()

        workPole = Pole(GRF[0],GRF[1],'xox','xoxx')
        
        #--------------------------------------------
        #Get screenshot from openreach website - NOTE AS OF V2.0r210222 THIS SECTION HAS BEEN OVERHAULED TO COMPLY WITH NEW OPENREACH LOGIN SYSTEM (AZURE SSO)
        
        time.sleep(5)

        threelines = driver.find_element_by_xpath("//*[@id='searchpanel']/div[1]/i")
        hover = ActionChains(driver).move_to_element(threelines).click(threelines)
        hover.perform()

        time.sleep(5)

        GRFdrop = driver.find_element_by_xpath("//*[@id='mat-expansion-panel-header-3']/span[1]/mat-panel-title")
        hover = ActionChains(driver).move_to_element(GRFdrop).click(GRFdrop)
        hover.perform()

        time.sleep(5)

        GRFxbox = driver.find_element_by_xpath("//*[@id='mat-input-2']")
        GRFybox = driver.find_element_by_xpath("//*[@id='mat-input-3']")

        GRFxbox.click
        GRFxbox.send_keys(workPole.xGRF)

        time.sleep(2)

        GRFybox.click
        GRFybox.send_keys(workPole.yGRF)

        time.sleep(2)

        GRFybox.send_keys(Keys.RETURN)

        time.sleep(5)

        zoomInButton = driver.find_element_by_xpath("//*[@id='map']/div[3]/div[2]/div[4]/button[1]")

        hover = ActionChains(driver).move_to_element(zoomInButton).click(zoomInButton)
        hover.perform()

        time.sleep(5)

        hover.perform()

        time.sleep(3)

        threelinesPost = driver.find_element_by_xpath("//*[@id='searchpanel']/div[1]/i")
        hover = ActionChains(driver).move_to_element(threelinesPost).click(threelinesPost)
        hover.perform()

        time.sleep(3)

        downloadButton = driver.find_element_by_xpath("//*[@id='cdk-accordion-child-16']/div/div/div[1]/i")
        hover = ActionChains(driver).move_to_element(downloadButton).click(downloadButton)
        hover.perform()

        time.sleep(2)

        CME = driver.find_element_by_xpath("//*[@id='mat-dialog-0']/app-download-map-dialog/div[2]/div/button[2]/span")
        hover = ActionChains(driver).move_to_element(CME).click(CME)
        hover.perform()
        time.sleep(3)
                
        
        #--------------------------------------------
        #Perform file operations for screenshots and csv file
        folder_path = (r'C:\Users\%s\Downloads\*' %user) #navigate to downloads

        files = glob.glob(folder_path)
        
        max_file = max(files, key=os.path.getctime) #get path of most recent file in downloads folder
    
        image = cv2.imread(max_file) #Read screenshot
        
        height, width, channels = image.shape #Get image dimensions
        
        window_name = 'Image' #name of image window
        
        center_coordinates = (int(width/2),int(height/2)) #coords of picture center

        radius = 30 #circle radius in pixels

        color = (0, 0, 255) #(Blue,Green,Red)

        thickness = 5 #circle thickness in pixels

        cv2.circle(image, center_coordinates, radius, color, thickness) #Draw circle on image depending on above parameters

        cv2.imwrite(screenshotPath + "\\" + filename.removesuffix('.xlsx') + '.png', image)
        
        img = drawing.image.Image(screenshotPath + "\\" + filename.removesuffix('.xlsx') + '.png')

        cover.add_image(img,'B6')

        A55Book.save(A55Path + "\\" + filename) #Save final xlsx file

        
         
        continue
    else:
        continue



